import tkinter as tk
from tkinter import filedialog, scrolledtext
import threading
import sys
import os
import requests
import pandas as pd
from base64 import b64encode
from datetime import datetime, timedelta
from openpyxl import load_workbook
import os
import time
# Save your original script in a file, say script_main.py
# import file as main_script
import tkinter as tk
from tkinter import filedialog, scrolledtext
import threading
import sys
import os
import requests
import pandas as pd
from base64 import b64encode
from datetime import datetime, timedelta
from openpyxl import load_workbook
import os
import time
# Save your original script in a file, say script_main.py
stop_button = False

def main(pat1,member_id1,input_path1):
    """
    Main function to analyze performance data from Azure DevOps
    based on a list of users provided in an Excel file.

    Args:
        pat1 (str): Personal Access Token for Azure DevOps.
        member_id1 (str): Azure DevOps member ID.
        input_path1 (str): Path to Excel file with input data.
    """

    # -------------- CONFIG --------------
    pat = pat1
    member_id = member_id1
    input_file = input_path1  # Excel file with 'emails' and 'days' columns
    output_file = "performance.xlsx"
    # ------------------------------------
    
    def format_allocation(value):
        """
        Format allocation value as a percentage or return 'N/A'.

        Args:
            value (float|int|None): The allocation value.

        Returns:
            str: Formatted percentage string or 'N/A'.
        """
        if pd.isna(value):
            return "N/A"
        if isinstance(value, (float, int)):
            return f"{value * 100:.0f}%" if value <= 1 else f"{value:.0f}%"
        return str(value)

    def format_date(value):
        """
        Format a pandas Timestamp or return 'N/A' for invalid values.

        Args:
            value (datetime|str|None): The date value.

        Returns:
            str: Date in MM/DD/YYYY format or 'N/A'.
        """
        if pd.isna(value) or not value:
            return "N/A"
        if isinstance(value, pd.Timestamp):
            return value.strftime('%m/%d/%Y')
        return str(value)

    # Retry wrapper
    def make_request(method, url, headers=None, json=None, max_retries=3, timeout=10):
        """
        Make a HTTP request with retry mechanism.

        Args:
            method (str): HTTP method ('get' or 'post').
            url (str): Target URL.
            headers (dict): HTTP headers.
            json (dict): Request payload for POST.
            max_retries (int): Retry attempts.
            timeout (int): Timeout per request.

        Returns:
            requests.Response or None: HTTP response object or None on failure.
        """
        for attempt in range(max_retries):
            try:
                if method.lower() == 'get':
                    return requests.get(url, headers=headers, timeout=timeout)
                elif method.lower() == 'post':
                    return requests.post(url, headers=headers, json=json, timeout=timeout)
            except requests.exceptions.RequestException as e:
                print(f"⚠️ Network error on attempt {attempt + 1}/{max_retries}: {e}")
                time.sleep(1)
        print(f"❌ Failed after {max_retries} attempts: {url}")
        return None

    # Read input file safely
    try:
        df_input = pd.read_excel(input_file)
    except Exception as e:
        print(f"❌ Failed to read input file '{input_file}': {e}")
        exit(1)

    auth = b64encode(f':{pat}'.encode()).decode()
    headers = {
        'Authorization': f'Basic {auth}',
        'Content-Type': 'application/json'
    }

    # Fetch organization list
    orgs_url = f"https://app.vssps.visualstudio.com/_apis/accounts?memberId={member_id}&api-version=7.0"
    orgs_response = make_request('get', orgs_url, headers=headers)
    time.sleep(0.5)
    if orgs_response is None or orgs_response.status_code != 200:
        print(f"❌ Failed to fetch organizations: {orgs_response.text if orgs_response else 'No response'}")
        exit(1)

    organizations = [org['accountName'] for org in orgs_response.json().get('value', [])]
    print(f"\n🔎 Found {len(organizations)} organizations.")

    for idx, row in df_input.iterrows():
        global stop_button
        if stop_button==True:
            break
        excel_id = row['ID']
        email = row['EmailID']
        num_days = int(row['Days'])
        department = row.get('Department', 'N/A');
        designation = row.get('Designation', 'N/A');
        client = row.get('Client', 'N/A');
        project_excel = row.get('Project', 'N/A');
        Role = row.get('Role', 'N/A');
        function = row.get('Function', 'N/A');
        manager = row.get('Manager', 'N/A');
        onwer = row.get('BU Owner', 'N/A');
        allocation = format_allocation(row.get('Allocation', 'N/A'))
        date_of_joining = format_date(row.get('Date of Joining', 'N/A'))

        allocation_status = row.get('Allocation Status', 'N/A');
        project_code = row.get('Project Code', 'N/A');
        internal_external = row.get('Internal/External', 'N/A');
        # date_of_joining = row.get('Date of Joining', 'N/A');
        working_status = format_date(row.get('Working Status', 'N/A'));
        experience_in_sht  = row.get('Experience in SHT', 'N/A');
        experience_before_sht = row.get('Experience before SHT', 'N/A');


        total_done = coding = design = review = other = ftar_sum = 0
        full_name = None

        print(f"\n📌 Searching for user: {email} for past {num_days} days")
        today = datetime.utcnow()
        start_date = (today - timedelta(days=num_days)).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

        for org in organizations:
            print(f"\n🌐 Organization: {org}")
            projects_url = f"https://dev.azure.com/{org}/_apis/projects?api-version=7.0"
            projects_response = make_request('get', projects_url, headers=headers)
            time.sleep(0.5)

            if projects_response is None or projects_response.status_code != 200:
                print(f"❌ Failed to fetch projects for org: {org}")
                continue

            projects = projects_response.json().get('value', [])
            print(f"🔧 Found {len(projects)} projects in {org}")

            for project in projects:
                project_name = project['name']
                print(f"   🔍 Searching project: {project_name}")

                wiql = {
                    "query": f"""
                        SELECT [System.Id]
                        FROM WorkItems
                        WHERE
                            [System.AssignedTo] CONTAINS '{email}'
                            AND [System.WorkItemType] IN ('Task')
                            AND [System.TeamProject] = '{project_name}'
                            AND [System.ChangedDate] >= '{start_date}'
                            AND [System.ChangedDate] <= '{end_date}'
                            AND [System.State] = 'Done'
                        ORDER BY [System.ChangedDate] DESC
                    """
                }

                wiql_url = f"https://dev.azure.com/{org}/{project_name}/_apis/wit/wiql?api-version=7.0"
                wiql_response = make_request('post', wiql_url, headers=headers, json=wiql)
                time.sleep(0.5)

                if wiql_response is None or wiql_response.status_code != 200:
                    print(f"      ⚠️ WIQL query failed for {project_name}")
                    continue

                work_items = wiql_response.json().get('workItems', [])
                if not work_items:
                    print(f"      📭 No done work items found.")
                    continue

                ids = ','.join(str(wi['id']) for wi in work_items)
                details_url = (
                    f"https://dev.azure.com/{org}/_apis/wit/workitems?ids={ids}&fields=" +
                    ",".join([
                        "System.WorkItemType",
                        "System.State",
                        "System.Tags",
                        "Microsoft.VSTS.CMMI.TaskType",
                        "Custom.FTARValue",
                        "Custom.TicketType",
                        "System.AssignedTo"
                    ]) + "&api-version=7.0"
                )
                details_response = make_request('get', details_url, headers=headers)
                time.sleep(0.5)

                if details_response is None or details_response.status_code != 200:
                    print(f"      ❌ Failed to fetch work item details.")
                    continue

                for item in details_response.json()['value']:
                    fields = item.get("fields", {})
                    task_type = fields.get("Microsoft.VSTS.CMMI.TaskType", "").strip() or fields.get("Custom.TicketType", "").strip()
                    ftar_value = fields.get("Custom.FTARValue")

                    assigned_to = fields.get("System.AssignedTo", {})
                    if isinstance(assigned_to, dict):
                        full_name = assigned_to.get("displayName", email)

                    total_done += 1
                    if task_type == "Code Review":
                        review += 1
                    elif task_type == "Design":
                        design += 1
                    elif task_type in ["Coding/Implementation", "Coding"]:
                        coding += 1
                        if ftar_value == 1:
                            ftar_sum += 1
                    else:
                        other += 1

                    print(f"      ✅ Task ID: {item['id']} - Project: {project_name} - Task Type: {task_type or 'N/A'}")

        result_row = {
            "ID": excel_id,
            "Name/Email id": email,
            "Full Name": full_name or email,
            "Number of days (X)": num_days,
            "Number of tickets in last X days": total_done,
            "Coding": coding,
            "Design": design,
            "Code review": review,
            "Others": other,
            "If Coding, then average of FTAR": ftar_sum / coding if coding else 0,
            "Department": department,
            "Designation": designation,
            "Client": client,
            "Project": project_excel,
            "Role": Role,
            "Function": function,
            "Manager": manager,
            "BU Owner": onwer,
            "Allocation": allocation,
            "Allocation Status": allocation_status,
            "Project Code": project_code,
            "Internal/External": internal_external,
            "Date of Joining": date_of_joining,
            "Working Status": working_status,
            "Experience in SHT": experience_in_sht,
            "Experience before SHT": experience_before_sht
        }

        # -----------------
        # department = row.get('', 'N/A');
        # designation = row.get('Designation', 'N/A');
        # client = row.get('Client', 'N/A');
        # project = row.get('Project', 'N/A');
        # Role = row.get('Role', 'N/A');
        # function = row.get('Function', 'N/A');
        # manager = row.get('Manager', 'N/A');
        # onwer = row.get('BU Owner', 'N/A');
        # allocation = row.get('Allocation', 'N/A');
        # allocation_status = row.get('Allocation Status', 'N/A');
        # project_code = row.get('Project Code', 'N/A');
        # internal_external = row.get('Internal/External', 'N/A');
        # date_of_joining = row.get('Date of Joining', 'N/A');
        # working_status = row.get('Working Status', 'N/A');
        # experience_in_sht  = row.get('Experience in SHT', 'N/A');
        # experience_before_sht = row.get('Experience before SHT', 'N/A');
        # --------

        df_output = pd.DataFrame([result_row])

        # Append or write to performance.xlsx
        if os.path.exists(output_file):
            with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                workbook = load_workbook(output_file)
                sheet = workbook.active
                start_row = sheet.max_row
                df_output.to_excel(writer, index=False, header=False, startrow=start_row)
        else:
            df_output.to_excel(output_file, index=False)

    print("\n✅ Done. Data appended to 'performance.xlsx'")

# Redirect stdout to UI
class RedirectOutput:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, string):
        self.text_widget.insert(tk.END, string)
        self.text_widget.see(tk.END)

    def flush(self):
        pass

def run_script():
    global stop_button
    stop_button = False 
    pat = pat_entry.get()
    member_id = member_id_entry.get()
    input_path = input_file_path.get()

    if not pat or not member_id or not input_path:
        output_text.insert(tk.END, "❌ Please fill in all fields and upload the input file.\n")
        return

    # Set variables inside the script dynamically
    pat1234 = pat
    # print(pat)
    member_id1234 = member_id
    input_file1234 = input_path

    def target():
        
        main(pat,member_id,input_path)  # Call a main function that runs the logic

    global thread
    thread = threading.Thread(target=target)
    thread.start()

def stop_script():
    global stop_button
    stop_button = True

def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        input_file_path.set(file_path)

# GUI setup
root = tk.Tk()
root.title("Azure DevOps Performance Tracker")

tk.Label(root, text="PAT:").grid(row=0, column=0, sticky="e")
pat_entry = tk.Entry(root, width=60)
pat_entry.grid(row=0, column=1, padx=5, pady=5)

tk.Label(root, text="Member ID:").grid(row=1, column=0, sticky="e")
member_id_entry = tk.Entry(root, width=60)
member_id_entry.grid(row=1, column=1, padx=5, pady=5)

tk.Label(root, text="Input Excel File:").grid(row=2, column=0, sticky="e")
input_file_path = tk.StringVar()
tk.Entry(root, textvariable=input_file_path, width=45).grid(row=2, column=1, sticky="w")
tk.Button(root, text="Browse", command=browse_file).grid(row=2, column=2)

tk.Button(root, text="Run", command=run_script, bg="lightgreen").grid(row=3, column=1, sticky="w", pady=10)
tk.Button(root, text="Stop", command=stop_script, bg="lightcoral").grid(row=3, column=1, sticky="e", pady=10)

output_text = scrolledtext.ScrolledText(root, width=100, height=30)
output_text.grid(row=4, column=0, columnspan=3, padx=10, pady=10)

# Redirect print output
sys.stdout = RedirectOutput(output_text)
sys.stderr = RedirectOutput(output_text)

thread = None
root.mainloop()

# Redirect stdout to UI
class RedirectOutput:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, string):
        self.text_widget.insert(tk.END, string)
        self.text_widget.see(tk.END)

    def flush(self):
        pass

def run_script():
    pat = pat_entry.get()
    member_id = member_id_entry.get()
    input_path = input_file_path.get()

    if not pat or not member_id or not input_path:
        output_text.insert(tk.END, "❌ Please fill in all fields and upload the input file.\n")
        return

    # Set variables inside the script dynamically
    pat = pat
    member_id = member_id
    
    input_file = input_path

    def target():
        main()  # Call a main function that runs the logic

    global thread
    thread = threading.Thread(target=target)
    thread.start()

def stop_script():
    global thread
    if thread and thread.is_alive():
        output_text.insert(tk.END, "⚠️ Stop requested (cannot force stop due to Python threading limitations).\n")

def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        input_file_path.set(file_path)

# GUI setup
root = tk.Tk()
root.title("Azure DevOps Performance Tracker")

tk.Label(root, text="PAT:").grid(row=0, column=0, sticky="e")
pat_entry = tk.Entry(root, width=60)
pat_entry.grid(row=0, column=1, padx=5, pady=5)

tk.Label(root, text="Member ID:").grid(row=1, column=0, sticky="e")
member_id_entry = tk.Entry(root, width=60)
member_id_entry.grid(row=1, column=1, padx=5, pady=5)

tk.Label(root, text="Input Excel File:").grid(row=2, column=0, sticky="e")
input_file_path = tk.StringVar()
tk.Entry(root, textvariable=input_file_path, width=45).grid(row=2, column=1, sticky="w")
tk.Button(root, text="Browse", command=browse_file).grid(row=2, column=2)

tk.Button(root, text="Run", command=run_script, bg="lightgreen").grid(row=3, column=1, sticky="w", pady=10)
tk.Button(root, text="Stop", command=stop_script, bg="lightcoral").grid(row=3, column=1, sticky="e", pady=10)

output_text = scrolledtext.ScrolledText(root, width=100, height=30)
output_text.grid(row=4, column=0, columnspan=3, padx=10, pady=10)

# Redirect print output
sys.stdout = RedirectOutput(output_text)
sys.stderr = RedirectOutput(output_text)

thread = None
root.mainloop()